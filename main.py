# import openpyxl
from openpyxl import workbook, load_workbook
wb = load_workbook('Grades.xlsx')
ws = wb.active
print(ws["A3"].value)